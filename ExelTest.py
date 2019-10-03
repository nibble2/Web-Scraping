# 엑셀에 스크랩핑한 데이터 조작하기

from openpyxl import load_workbook

work_book = load_workbook('prac.xlsx')
work_sheet = work_book['prac']

# 데이터를 읽어봅니다.
temp_text = work_sheet.cell(row = 2, column = 2).value
print(temp_text)

temp_text2 = work_sheet.cell(row=3, column= 2, value='이다나')

# 수정한 엑셀파일을 저장합니다.
# 참고: 다른이름으로 저장할 수도 있습니다.
work_book.save('prac.xlsx')
# work_book.save('prac02.xlsx')