# 네이버 영화 순위 / 제목 / 평점을 스크랩핑 하여 엑셀파일에 저장
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

work_book = load_workbook('prac.xlsx')
work_sheet = work_book['movie']

# URL을 읽어서 HTML를 받아오고,
headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://movie.naver.com/movie/sdb/rank/rmovie.nhn?sel=pnt&date=20190909',headers=headers)

soup = BeautifulSoup(data.text, 'html.parser')

# tr 불러오기
movies = soup.select('#old_content > table > tbody > tr')


rank = 1
row = 2

for movie in movies:
    a_tag = movie.select_one('td.title > div > a')
    if a_tag is not None:
        title = a_tag.text
        star = movie.select_one('td.point').text

        work_sheet.cell(row, column=1, value=rank)
        work_sheet.cell(row, column=2, value=title)
        work_sheet.cell(row, column=3, value=star)
        row += 1
        rank += 1

work_book.save('prac.xlsx')
